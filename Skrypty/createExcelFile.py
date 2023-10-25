import openpyxl as openpyxl
import sympy
import os
bestPath = r'C:\Users\dkmak\OneDrive\Pulpit\Generyczne\TinyGP-Java-master\best.txt'
probPath = r'C:\Users\dkmak\OneDrive\Pulpit\Generyczne\TinyGP-Java-master\problem.dat'
def refactorData():
    with open(bestPath, 'r') as file:
        best = file.read()
        best = best.replace('.', ',')
        best = best.replace('X1', 'A1')
        best = best.replace('X2', 'A2')
        with open(bestPath, 'w') as file:
            file.write(best)
def writeToExcel():
    workbook = openpyxl.load_workbook("Wykresy.xlsx")
    worksheet = workbook.active
    data = []
    with open(probPath, 'r') as file:
        lines = file.readlines()
        for line in lines[1:]:
            data.append(line.strip().split(' '))
    with open(bestPath, 'r') as file:
        best = file.read()
        for i, row in enumerate(data):
            worksheet.cell(row=i+1, column=1, value=row[0].replace('.', ','))
            value_as_number = float(row[3])
            worksheet.cell(row=i + 1, column=3, value=value_as_number)
            toSave = best.replace('A1', str(row[0])).replace('A2', str(row[0]))
            newValue = (sympy.sympify(toSave.replace(',', '.')))
            worksheet.cell(row=i + 1, column=2, value=float(newValue))
    fileName = 'example.xlsx'
    i=0
    while os.path.exists(fileName):
        fileName = 'example' + str(i) + '.xlsx'
        i+=1
    workbook.save(fileName)
    workbook.close()
if __name__ == '__main__':
    refactorData()
    writeToExcel()